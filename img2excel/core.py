"""
核心功能模块 - ImageToExcel类
"""

import os
from typing import Tuple, Optional, Union
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from .utils import resize_image, rgb_to_hex


class ImageToExcel:
    """
    将图片转换为Excel像素画的主要类
    
    支持自定义单元格尺寸、保持原比例、批量处理等功能
    """
    
    def __init__(self, image_path: str):
        """
        初始化ImageToExcel实例
        
        Args:
            image_path: 图片文件路径
        """
        self.image_path = image_path
        self.image = None
        self.workbook = None
        self.worksheet = None
        
        # 验证图片文件
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"图片文件不存在: {image_path}")
        
        # 加载图片
        self._load_image()
    
    def _load_image(self):
        """加载图片文件"""
        try:
            self.image = Image.open(self.image_path)
            # 转换为RGB模式（处理RGBA等格式）
            if self.image.mode != 'RGB':
                self.image = self.image.convert('RGB')
        except Exception as e:
            raise ValueError(f"无法加载图片文件: {e}")
    
    def convert_to_excel(
        self,
        output_path: str,
        cell_width: Optional[int] = None,
        cell_height: Optional[int] = None,
        max_width: Optional[int] = None,
        max_height: Optional[int] = None,
        keep_ratio: bool = True,
        sheet_name: str = "PixelArt"
    ) -> str:
        """
        将图片转换为Excel文件
        
        Args:
            output_path: 输出Excel文件路径
            cell_width: 单元格宽度（像素）
            cell_height: 单元格高度（像素）
            max_width: 最大宽度（单元格数量）
            max_height: 最大高度（单元格数量）
            keep_ratio: 是否保持原比例
            sheet_name: 工作表名称
            
        Returns:
            输出文件路径
        """
        # 计算目标尺寸
        target_size = self._calculate_target_size(
            max_width, max_height, keep_ratio
        )
        
        # 调整图片尺寸
        resized_image = resize_image(self.image, target_size)
        
        # 创建Excel工作簿
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        
        # 设置单元格尺寸
        self._set_cell_dimensions(
            resized_image.size[0], 
            resized_image.size[1],
            cell_width,
            cell_height
        )
        
        # 渲染图片到Excel
        self._render_image_to_excel(resized_image)
        
        # 保存文件
        self.workbook.save(output_path)
        
        return output_path
    
    def _calculate_target_size(
        self, 
        max_width: Optional[int], 
        max_height: Optional[int], 
        keep_ratio: bool
    ) -> Tuple[int, int]:
        """
        计算目标图片尺寸
        
        Args:
            max_width: 最大宽度
            max_height: 最大高度
            keep_ratio: 是否保持比例
            
        Returns:
            (宽度, 高度) 元组
        """
        original_width, original_height = self.image.size
        
        if max_width is None and max_height is None:
            # 如果没有指定尺寸限制，使用原尺寸
            return original_width, original_height
        
        if keep_ratio:
            # 保持原比例
            if max_width is not None and max_height is not None:
                # 两个维度都指定了，选择较小的缩放比例
                scale_w = max_width / original_width
                scale_h = max_height / original_height
                scale = min(scale_w, scale_h)
            elif max_width is not None:
                # 只指定宽度
                scale = max_width / original_width
            else:
                # 只指定高度
                scale = max_height / original_height
            
            new_width = int(original_width * scale)
            new_height = int(original_height * scale)
        else:
            # 不保持比例，使用指定的尺寸
            new_width = max_width if max_width is not None else original_width
            new_height = max_height if max_height is not None else original_height
        
        return new_width, new_height
    
    def _set_cell_dimensions(
        self, 
        width: int, 
        height: int,
        cell_width: Optional[int] = None,
        cell_height: Optional[int] = None
    ):
        """
        设置Excel单元格的尺寸
        
        Args:
            width: 图片宽度（单元格数量）
            height: 图片高度（单元格数量）
            cell_width: 单元格宽度（像素）
            cell_height: 单元格高度（像素）
        """
        # 设置列宽
        for col in range(1, width + 1):
            col_letter = get_column_letter(col)
            if cell_width:
                self.worksheet.column_dimensions[col_letter].width = cell_width / 7  # openpyxl使用字符单位
            else:
                self.worksheet.column_dimensions[col_letter].width = 2  # 默认宽度
        
        # 设置行高
        for row in range(1, height + 1):
            if cell_height:
                self.worksheet.row_dimensions[row].height = cell_height * 0.75  # openpyxl使用磅为单位
            else:
                self.worksheet.row_dimensions[row].height = 15  # 默认高度
    
    def _render_image_to_excel(self, image: Image.Image):
        """
        将图片渲染到Excel中
        
        Args:
            image: PIL图片对象
        """
        width, height = image.size
        
        # 获取图片数据
        img_data = image.load()
        
        print(f"正在渲染图片到Excel... ({width}x{height})")
        
        # 逐像素设置单元格颜色
        for y in range(height):
            for x in range(width):
                # 获取像素RGB值
                r, g, b = img_data[x, y]
                
                # 转换为十六进制颜色
                hex_color = rgb_to_hex(r, g, b)
                
                # 创建填充样式
                fill = PatternFill(
                    start_color=hex_color,
                    end_color=hex_color,
                    fill_type="solid"
                )
                
                # 设置单元格填充
                cell = self.worksheet.cell(row=y+1, column=x+1)
                cell.fill = fill
        
        print("渲染完成！")
    
    def get_image_info(self) -> dict:
        """
        获取图片信息
        
        Returns:
            包含图片信息的字典
        """
        if self.image is None:
            return {}
        
        return {
            "path": self.image_path,
            "size": self.image.size,
            "mode": self.image.mode,
            "format": self.image.format,
            "width": self.image.size[0],
            "height": self.image.size[1]
        }
    
    def preview_resize(
        self, 
        max_width: Optional[int] = None, 
        max_height: Optional[int] = None, 
        keep_ratio: bool = True
    ) -> Tuple[int, int]:
        """
        预览调整后的尺寸（不实际调整图片）
        
        Args:
            max_width: 最大宽度
            max_height: 最大高度
            keep_ratio: 是否保持比例
            
        Returns:
            (宽度, 高度) 元组
        """
        return self._calculate_target_size(max_width, max_height, keep_ratio)
